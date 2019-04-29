import random
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font


class Spelertje:
    def __init__(self, name, position, goals, birthCat, effort, weight, length,birthDate):
        self.name = name
        self.position= position
        self.goals = goals
        self.birthCat = birthCat
        self.effort = effort
        self.weight = weight
        self.length = length
        self.birthDate= birthDate

    def returnArray(self):
        return [self.name, self.position, self.goals, self.birthCat, self.effort, self.weight, self.length,self.birthDate]

    def __str__(self):
        return str(self.name) + " - " + str(self.position) + " - " + str(self.goals) + " - " + \
               str(self.birthCat) + " - " + str(self.effort) + " - " + str(self.weight) + " - " + str(self.length + "-" + str(self.birthDate))

class Spelertjes:
    def __init__(self):
        self.spelertjes = []
        self.effort = {1:"zeer goed", 2 : "goed", 3:"goed",4:"matig"}
        self.spelertjesValues = []

    def addSpeler(self, spelertje):
        self.spelertjes.append(spelertje)

    def returnCat(self, date):
        day =  date.timetuple().tm_yday
        if day <= 90:
            return 1
        if day <= 181:
            return 2
        if day <= 273:
            return 3
        if day <= 365:
            return 4

    def readFile(self, fileName, sheetName):
        wb = load_workbook(fileName)
        ws = wb[sheetName]

        for row in ws.rows:
            args = [cell.value for cell in row]
            birthDate = self.generateBirth()
            spelertje = Spelertje(args[0], args[2], args[3], self.returnCat(birthDate),self.effort[self.returnCat(birthDate)], args[6],args[7], birthDate)
            self.spelertjesValues.append(spelertje.returnArray())
            self.addSpeler(spelertje)

        self.spelertjes.remove(self.spelertjes[0])
        self.spelertjesValues.remove(self.spelertjesValues[0])

    def writeFile(self, fileName, sheetName, saveFileName):
        wb = load_workbook(fileName)
        ws = wb[sheetName]

        font = Font(name='Calibri',size=12,bold=True)
        header = ["naam", "positie", "aantal gemaakte goalen", "geboortecategorie", "inzet", "gewicht", "lengte","geboortedatum"]

        #fill the headers with the right values
        for i in range(len(header)):
            cellref = ws.cell(1,i+1)
            cellref.value = header[i]
            cellref.font = font
            #clear last column
            cellref = ws.cell(1,i+2)
            cellref.value = None

        #fill gegevens with new generated values
        for i in range(len(self.spelertjesValues)):
            for j in range(len(self.spelertjesValues[i])):
                cellref = ws.cell(row=i+2, column=j+1)
                cellref.value = self.spelertjesValues[i][j]
                #clear last column
                cellref = ws.cell(row=i + 2, column=j + 2)
                cellref.value = None

        wb.save(saveFileName)

    def generateBirth(self):
        startdate = datetime.date(2011, 1, 1)
        date = startdate + datetime.timedelta(random.randint(1,365))
        return date
