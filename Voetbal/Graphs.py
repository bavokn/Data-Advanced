from openpyxl import load_workbook
from openpyxl.chart.series import Series
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from Voetbal.Spelertjes import Spelertjes
import operator
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
    marker,
    BarChart
)

class visual:

    def drawScatterChart(self, fileName, sheetName, saveFileName = None):

        if saveFileName is None:
            saveFileName = fileName

        wb = load_workbook(fileName)
        ws = wb['gegevens']

        chart = ScatterChart()
        chart.title = "Scatter Chart"
        chart.style = 13
        chart.x_axis.scaling.min = 19
        chart.x_axis.scaling.max = 31
        chart.y_axis.scaling.min = 110
        chart.y_axis.scaling.max = 140
        chart.x_axis.title = 'gewicht'
        chart.y_axis.title = 'lengte'

        chart.legend = None

        xvalues = Reference(ws, min_col=6, min_row=2, max_row=101)
        values = Reference(ws, min_col=7, min_row=2, max_row=101)
        #fill x and y, skip first
        x=[]
        y=[]
        iterrows = iter(ws.rows)
        next(iterrows)
        for row in iterrows:
            x.append(row[5].value)
            y.append(row[6].value)

        series = Series(values, xvalues)
        series.graphicalProperties.line.noFill = True

        series.marker = marker.Marker('circle', 5.2)

        chart.series.append(series)

        # Style the lines
        s1 = chart.series[0]
        s1.marker.symbol = "circle"
        s1.marker.graphicalProperties.solidFill = "4076A9"  # Marker filling
        s1.marker.graphicalProperties.line.solidFill = "4076A9"  # Marker outline

        s1.graphicalProperties.line.noFill = True

        ws = wb[sheetName]

        ws.add_chart(chart, "L7")

        wb.save(saveFileName)

        area = np.pi * 20
        plt.scatter(x, y, s=area, alpha=1)
        plt.xlabel("gewicht")
        plt.ylabel("lengte")
        plt.grid(True,alpha=0.5)
        plt.axis([19,31,110,140])
        plt.show()


    def drawBarChart(self, fileName, sheetName, saveFileName = None):
        spelertjes = Spelertjes()
        if saveFileName is None:
            saveFileName = fileName
        #read all the data using openpyxl and write data to grafiek tab
        wb = load_workbook(fileName)
        ws = wb['gegevens']


        goals = {"staart":{1:0,2:0,3:0,4:0}, "linkervleugel":{1:0,2:0,3:0,4:0},"rechtervleugel":{1:0,2:0,3:0,4:0},
                 "piloot":{1:0,2:0,3:0,4:0},"keeper":{1:0,2:0,3:0,4:0}}

        positions = ["staart","linkervleuger","rechtervleuger","piloot","keeper"]

        iterrows = iter(ws.rows)
        next(iterrows)

        for row in iterrows:
            position = goals[row[1].value]
            position[row[3].value] += row[2].value
            goals[row[1].value] = position
        ws = wb['grafiek']

        for i in range(2,6):
            cellref = ws.cell(1, i)
            cellref.value = i - 1

        for i in range(2,7):
            cellref = ws.cell(i, 1)
            cellref.value = positions[i-2]
        row = 2

        for i in goals.values():
            column = 2
            for j in i.values():
                cellref = ws.cell(row, column)
                cellref.value = j
                column += 1
            row += 1

        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "goals per position per birth cat"
        chart1.y_axis.title = 'goals'
        chart1.x_axis.title = 'position'

        data = Reference(ws, min_col=2, min_row=1, max_row=6, max_col=5)
        cats = Reference(ws, min_col=1, min_row=1 , max_row=6)
        chart1.add_data(data,titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        ws.add_chart(chart1, "C24")
        wb.save(saveFileName)

        pd.DataFrame(goals).plot(kind='bar')
        plt.xlabel("geboortecategorie")
        plt.ylabel("aantal gemaakte goals")
        plt.grid(True, alpha=0.5)
        plt.show()



    def averageAndModus(self, fileName,):
        # read all the data using openpyxl and write data to grafiek tab
        wb = load_workbook(fileName)
        ws = wb['gegevens']

        goals = {"staart": 0, "linkervleugel": 0,
                 "rechtervleugel": 0,
                 "piloot": 0, "keeper": 0}

        # count amount of players in cat | is always 20 but just in case it changes...
        goalsCounter = {"staart": 0, "linkervleugel": 0,
                 "rechtervleugel": 0,
                 "piloot": 0, "keeper": 0}

        averageGoals = {"staart": 0, "linkervleugel": 0,
                 "rechtervleugel": 0,
                 "piloot": 0, "keeper": 0}

        modus = {"staart":{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0}, "linkervleugel":{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0},"rechtervleugel":{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0},
                 "piloot":{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0},"keeper":{0:0,1:0,2:0,3:0,4:0,5:0,6:0,7:0,8:0}}

        iterrows = iter(ws.rows)
        next(iterrows)
        for row in iterrows:
            temp = modus[row[1].value]
            temp[row[2].value] += 1
            modus[row[1].value] = temp
            goalsCounter[row[1].value] += 1
            goals[row[1].value] += row[2].value


        for i in goals:
            averageGoals[i] = goals[i]/goalsCounter[i]
        #still need to write this to excel file
        print("modus : ")
        for pos in modus:
            print(str(pos)+ " :"+ str(max(modus[pos], key=modus[pos].get)))
        print()
        print("Average Goals : ")
        for goal in averageGoals:
            print(str(goal) + " : " + str(averageGoals[goal]))

    def calculateQuartileAndStd(self,fileName):
            # read all the data using openpyxl and write data to grafiek tab
        wb = load_workbook(fileName)
        ws = wb['gegevens']

        data = []
        #skip first row
        iterrows = iter(ws.rows)
        next(iterrows)
        for row in iterrows:
            data.append(row[5].value)
        #calculate std and quartile 1
        print("Kwartiel 1 : " + str(np.percentile(data,25)))
        print("standaard afwijking : " + str(np.std(data)))

    def drawBoxPlot(self,fileName):
        wb = load_workbook(fileName)
        ws = wb['gegevens']

        data = {"linkervleugel":0,"rechtervleugel":0,"piloot":0}
        iterrows = iter(ws.rows)
        next(iterrows)
        for row in iterrows:
            pos = row[1].value
            if pos == "piloot" or pos == "linkervleugel" or pos == "rechtervleugel":
                data[row[1].value] += row[2].value

        plt.boxplot(data.values(), 0, 'rs', 0)
        plt.title("Boxplot")
        plt.xlabel("aantal gemaakte goals")
        plt.show()




